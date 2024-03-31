import requests
import datetime
import certifi
import urllib3
import openpyxl
import os

urllib3.disable_warnings()


def get_data_from_site():
    """
    Получение данных с сайта и сохранение в формате xlsx
    :return:
    """

    url = 'https://reestrs.minjust.gov.ru/rest/registry/ac648356-fe24-e11a-ceb0-87718bb81ed4/export?'

    date_now = str(datetime.datetime.now().date())
    # print(date_now)

    headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 '
                             'Safari/537.36'}

    response = requests.get(url, verify=False)
    # response = requests.get(url, headers=headers, timeout=5, verify=certifi.where())
    with open(f'./src/kazaki {date_now}.xlsx', 'wb') as output:
        output.write(response.content)


def reading_xlsx_file():
    """
    Чтение файла в формате xlsx
    :return:
    """

    # считываем данные из файла
    work_book = openpyxl.load_workbook("./src/test_kazaki_1.xlsx")

    # получаем активный лист
    work_sheet = work_book.active

    # создаем список для хранения данных
    data_list = []

    # перебираем строки
    for line in range(1, work_sheet.max_row + 1):

        # проверяем параметры добавления данных в список
        if work_sheet[f'A{line}'].value == 'Реестр казачьих обществ':

            new_line = []  # создаем новый список

            # считываем данные ячеек
            for col in work_sheet.iter_cols(1, work_sheet.max_column):

                # добавляем данные ячейки в список
                new_line.append(str(col[line - 1].value).replace('\xa0', ' '))

            # добавляем список с данными строки в общий список
            data_list.append(new_line)

    for i in data_list:
        print(i)


if __name__ == '__main__':
    reading_xlsx_file()

